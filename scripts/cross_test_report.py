#!/usr/bin/env python3
"""
Cross-Testing Report Generator for Standards Suite
Analyzes 19 standards files for cross-reference consistency,
severity alignment, broken links, and contradictions.
"""

import sys
import os

# RELATIVE paths — works standalone, as submodule, or in CI.
# Layout: <repo-root>/scripts/ and <repo-root>/upload/
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# xlsx skill helpers — these live in the sandbox under skills/xlsx/
# (not part of the corpus repo itself, but tolerated when present).
XLSX_SKILL_DIR = os.environ.get(
    'XLSX_SKILL_DIR',
    os.path.join(os.path.dirname(REPO_ROOT), 'skills', 'xlsx')
)
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# DATA: Standards Registry
# ============================================================

STANDARDS = {
    "STD-DOC-002": {
        "file": "MARKDOWN_STANDARD.md",
        "name": "Markdown Formatting",
        "version": "2.2.0",
        "level_declared": "[W] Warning",
        "level_in_id_system": "[W] Warning",
        "level_in_impl_order": "[W]",
    },
    "STD-DOC-003": {
        "file": "UNICODE_POLICY.md",
        "name": "No-Unicode Policy",
        "version": "2.1.3",
        "level_declared": "[C]+[W]+[I]",
        "level_in_id_system": "[C]+[W]+[I]",
        "level_in_impl_order": "[C]+[W]+[I]",
    },
    "STD-DOC-004": {
        "file": "README_TEMPLATE.md",
        "name": "README Template",
        "version": "2.1",
        "level_declared": "[W] Warning",
        "level_in_id_system": "(not in registry)",
        "level_in_impl_order": "(not in table)",
    },
    "STD-DOC-005": {
        "file": "CODE_EXAMPLES_GUIDE.md",
        "name": "Code Examples Guide",
        "version": "1.1",
        "level_declared": "[W] Warning",
        "level_in_id_system": "[W] Warning",
        "level_in_impl_order": "[W]",
    },
    "STD-FE-001": {
        "file": "FRONTEND_STANDARD.md",
        "name": "Frontend Standard",
        "version": "1.5",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-GIT-001": {
        "file": "GITHUB_STANDARD.md",
        "name": "GitHub Core Standard",
        "version": "2.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-GIT-002": {
        "file": "GITHUB_SANDBOX_STANDARD.md",
        "name": "GitHub Sandbox Safety",
        "version": "1.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-A11Y-001": {
        "file": "WCAG_2.1_AA_STANDARD.md",
        "name": "WCAG 2.1 AA Accessibility",
        "version": "1.1",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical (DEPRECATED)",
        "level_in_impl_order": "[C]",
    },
    "STD-TEST-001": {
        "file": "TESTING_STANDARD.md",
        "name": "Testing Standard",
        "version": "1.1",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-ERR-001": {
        "file": "ERROR_HANDLING_STANDARD.md",
        "name": "Error Handling Core",
        "version": "2.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-ERR-002": {
        "file": "ERROR_RECOVERY_STANDARD.md",
        "name": "Error Recovery",
        "version": "1.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-SEC-001": {
        "file": "SECURITY_STANDARD.md",
        "name": "Security Core",
        "version": "2.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-SEC-002": {
        "file": "SECURITY_EXTENDED_STANDARD.md",
        "name": "Security Extended",
        "version": "1.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-ENV-001": {
        "file": "REPRODUCIBILITY-STANDARD.md",
        "name": "Reproducibility",
        "version": "2.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-ENV-002": {
        "file": "ZAI_INTEGRATION_STANDARD.md",
        "name": "Z.ai Integration",
        "version": "1.1",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-ARCH-001": {
        "file": "IMPLEMENTATION_ORDER.md",
        "name": "Implementation Order",
        "version": "2.2",
        "level_declared": "[W] Warning",
        "level_in_id_system": "[W] Warning",
        "level_in_impl_order": "[W]",
    },
    "STD-META-001": {
        "file": "STANDARD_ID_SYSTEM.md",
        "name": "Standard ID System",
        "version": "1.1",
        "level_declared": "[W] Warning",
        "level_in_id_system": "[W] Warning",
        "level_in_impl_order": "[W]",
    },
    "STD-AGENT-001": {
        "file": "SUBAGENT_STANDARD.md",
        "name": "Subagent Standard",
        "version": "1.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
    "STD-AGENT-002": {
        "file": "ORCHESTRATION_STANDARD.md",
        "name": "Orchestration Standard",
        "version": "1.0",
        "level_declared": "[C] Critical",
        "level_in_id_system": "[C] Critical",
        "level_in_impl_order": "[C]",
    },
}

# ============================================================
# DATA: Cross-references (A -> B means A references B)
# ============================================================

CROSS_REFS = {
    "STD-DOC-002": ["STD-DOC-003"],
    "STD-DOC-003": ["STD-DOC-002"],
    "STD-DOC-004": ["STD-DOC-002"],
    "STD-DOC-005": ["STD-DOC-002", "STD-DOC-003", "STD-A11Y-001", "STD-SEC-001"],
    "STD-ARCH-001": [
        "STD-DOC-002", "STD-DOC-003", "STD-DOC-005", "STD-FE-001",
        "STD-GIT-001", "STD-GIT-002", "STD-A11Y-001", "STD-TEST-001",
        "STD-ERR-001", "STD-ERR-002", "STD-SEC-001", "STD-SEC-002",
        "STD-ENV-001", "STD-ENV-002", "STD-AGENT-001", "STD-AGENT-002",
        "STD-META-001",
    ],
    "STD-META-001": ["STD-ARCH-001", "STD-DOC-002"],
    "STD-FE-001": ["STD-A11Y-001", "STD-ERR-001", "STD-SEC-001", "STD-GIT-001", "STD-ENV-001"],
    "STD-GIT-001": ["STD-GIT-002", "STD-AGENT-001", "STD-FE-001"],
    "STD-GIT-002": ["STD-GIT-001", "STD-ENV-002", "STD-AGENT-001", "STD-FE-001"],
    "STD-A11Y-001": ["STD-FE-001", "STD-TEST-001", "STD-DOC-005"],
    "STD-TEST-001": ["STD-ERR-001", "STD-ERR-002", "STD-SEC-001", "STD-ENV-001", "STD-ENV-002"],
    "STD-ERR-001": ["STD-ERR-002", "STD-SEC-001", "STD-AGENT-001", "STD-AGENT-002"],
    "STD-ERR-002": ["STD-ERR-001", "STD-AGENT-002", "STD-SEC-001", "STD-GIT-002"],
    "STD-SEC-001": ["STD-SEC-002", "STD-ENV-001", "STD-ERR-001"],
    "STD-SEC-002": ["STD-SEC-001", "STD-ERR-001", "STD-A11Y-001"],
    "STD-ENV-001": ["STD-ERR-001", "STD-ERR-002", "STD-SEC-001", "STD-GIT-001", "STD-FE-001", "STD-ENV-002", "STD-TEST-001"],
    "STD-ENV-002": ["STD-ENV-001"],
    "STD-AGENT-001": [
        "STD-AGENT-002", "STD-ENV-002", "STD-GIT-001", "STD-GIT-002",
        "STD-ERR-001", "STD-META-001", "STD-DOC-003", "STD-DOC-002", "STD-ENV-001",
    ],
    "STD-AGENT-002": ["STD-AGENT-001", "STD-GIT-001", "STD-GIT-002", "STD-ERR-001", "STD-ENV-002", "STD-ARCH-001"],
}

# ============================================================
# DATA: Issues Found
# ============================================================

SEVERITY_ISSUES = [
    {
        "id": "SEV-001",
        "type": "Severity Mismatch",
        "source": "STD-DOC-002",
        "section": "Header / Section 2",
        "description": "MARKDOWN_STANDARD declares [W] Warning but references emoji/unicode rules from STD-DOC-003 which are [C] Critical. Same prohibition rule should have same severity.",
        "severity": "HIGH",
        "recommendation": "Upgrade STD-DOC-002 to [C] Critical for character prohibition rules, or clarify that [W] applies only to formatting (not content policy).",
    },
    {
        "id": "SEV-002",
        "type": "Severity Mismatch",
        "source": "STD-DOC-003",
        "section": "Section 3 / Level Application table",
        "description": "UNICODE_POLICY says documentation/README is [W] Warning, but same emoji prohibition is [C] in production code. Same rule, different severity creates enforcement gap.",
        "severity": "HIGH",
        "recommendation": "Unify to [C] for emoji/unicode prohibition across all contexts, or explicitly document why [W] is acceptable for docs.",
    },
    {
        "id": "SEV-003",
        "type": "Severity Mismatch",
        "source": "STD-A11Y-001",
        "section": "STD-META-001 registry",
        "description": "WCAG standard is marked as [C] Critical in its own header but STD-META-001 registry shows it as [C] Critical (DEPRECATED). If deprecated, references from STD-FE-001, STD-DOC-005, STD-SEC-002 may be stale.",
        "severity": "MEDIUM",
        "recommendation": "Clarify: is STD-A11Y-001 active or deprecated? If deprecated, what replaces it? Update all references.",
    },
    {
        "id": "SEV-004",
        "type": "Registry Gap",
        "source": "STD-DOC-004",
        "section": "STD-META-001 registry / STD-ARCH-001 table",
        "description": "README_TEMPLATE (STD-DOC-004) is not listed in STD-META-001 registry and not in STD-ARCH-001 implementation table. Other standards reference it (STD-DOC-002, STD-DOC-003).",
        "severity": "MEDIUM",
        "recommendation": "Add STD-DOC-004 to both STD-META-001 registry and STD-ARCH-001 implementation table.",
    },
]

BROKEN_REFS = [
    {
        "id": "REF-001",
        "type": "Stale Version Reference",
        "source": "STD-DOC-002",
        "section": "Header Related",
        "target": "STD-DOC-003",
        "detail": "References 'No-Unicode Policy v2.1' but current version is v2.1.3. Should specify at least v2.1.",
        "severity": "LOW",
    },
    {
        "id": "REF-002",
        "type": "Stale Version Reference",
        "source": "STD-DOC-003",
        "section": "Section 8.3",
        "detail": "References 'MARKDOWN_STANDARD v2.1' but current version is v2.2.0. Cross-reference version is outdated.",
        "severity": "LOW",
    },
    {
        "id": "REF-003",
        "type": "Stale Version Reference",
        "source": "STD-DOC-004",
        "section": "Section 2 Note",
        "detail": "References 'MARKDOWN_STANDARD v2.1 section 7' but current version is v2.2.0.",
        "severity": "LOW",
    },
    {
        "id": "REF-004",
        "type": "Stale Version Reference",
        "source": "STD-ENV-002",
        "section": "Footer",
        "detail": "States 'complies with MARKDOWN_STANDARD v2.1 (level [W])' but current version is v2.2.0.",
        "severity": "LOW",
    },
    {
        "id": "REF-005",
        "type": "Stale Version Reference",
        "source": "STD-ENV-002",
        "section": "Section 3.1",
        "detail": "References 'REPRODUCIBILITY-STANDARD (STD-ENV-001 v1.1)' but current version is v2.0.",
        "severity": "MEDIUM",
    },
    {
        "id": "REF-006",
        "type": "Missing Back-Reference",
        "source": "STD-SEC-001",
        "section": "Cross-refs section",
        "target": "STD-FE-001",
        "detail": "STD-FE-001 references STD-SEC-001 (Section 10.3, 10.4) but STD-SEC-001 cross-refs do not list STD-FE-001. Asymmetric link.",
        "severity": "MEDIUM",
    },
    {
        "id": "REF-007",
        "type": "Missing Back-Reference",
        "source": "STD-ENV-001",
        "section": "Cross-refs section",
        "target": "STD-AGENT-001",
        "detail": "STD-AGENT-001 references STD-ENV-001 but STD-ENV-001 cross-refs do not list STD-AGENT-001.",
        "severity": "MEDIUM",
    },
    {
        "id": "REF-008",
        "type": "Missing Back-Reference",
        "source": "STD-ENV-002",
        "section": "Cross-refs section",
        "target": "STD-AGENT-001, STD-AGENT-002, STD-GIT-002",
        "detail": "STD-AGENT-001, STD-AGENT-002, STD-GIT-002 all reference STD-ENV-002 but STD-ENV-002 cross-refs only list STD-ENV-001. Multiple missing back-refs.",
        "severity": "HIGH",
    },
    {
        "id": "REF-009",
        "type": "Missing Back-Reference",
        "source": "STD-FE-001",
        "section": "Cross-refs section",
        "target": "STD-GIT-002",
        "detail": "STD-GIT-002 references STD-FE-001 but STD-FE-001 cross-refs do not list STD-GIT-002.",
        "severity": "LOW",
    },
    {
        "id": "REF-010",
        "type": "Missing Cross-Ref Section",
        "source": "STD-DOC-005",
        "section": "Section 16",
        "target": "STD-DOC-003",
        "detail": "CODE_EXAMPLES_GUIDE references STD-DOC-003 in its cross-refs section (Section 16) but STD-DOC-003 does not reference STD-DOC-005 back.",
        "severity": "LOW",
    },
]

CONTRADICTIONS = [
    {
        "id": "CTR-001",
        "type": "Severity Contradiction",
        "sources": "STD-DOC-002 vs STD-DOC-003",
        "description": "STD-DOC-002 says documentation violations are [W] Warning (does not block merge). STD-DOC-003 says same emoji violations are [C] Critical in code. But the prohibition is the same rule. If emoji are truly prohibited, they should be [C] everywhere.",
        "impact": "HIGH",
        "resolution": "Unify: upgrade documentation character rules to [C], or explicitly document why [W] is intentional for docs.",
    },
    {
        "id": "CTR-002",
        "type": "Deprecation Ambiguity",
        "sources": "STD-A11Y-001 vs STD-META-001",
        "description": "STD-A11Y-001 header declares itself as [C] Critical and active. STD-META-001 registry marks it as [C] Critical (DEPRECATED). Is it active or not? 3 other standards still reference it (STD-FE-001, STD-DOC-005, STD-SEC-002).",
        "impact": "HIGH",
        "resolution": "Clarify deprecation status. If deprecated, add replacement standard. If active, remove DEPRECATED from registry.",
    },
    {
        "id": "CTR-003",
        "type": "Registry Completeness",
        "sources": "STD-DOC-004 vs STD-META-001",
        "description": "STD-DOC-004 (README_TEMPLATE) exists as a file with an STD-ID but is missing from STD-META-001 registry and STD-ARCH-001 implementation table. Both STD-DOC-002 and STD-DOC-003 reference it.",
        "impact": "MEDIUM",
        "resolution": "Add STD-DOC-004 to STD-META-001 registry and STD-ARCH-001 implementation table.",
    },
    {
        "id": "CTR-004",
        "type": "Missing STD-DOC-001",
        "sources": "STD-META-001 registry",
        "description": "STD-META-001 registry lists STD-DOC-001 as [W] Warning (DEPRECATED) but no file exists for it. Is this an old standard that was removed? No replacement is documented.",
        "impact": "LOW",
        "resolution": "Document what STD-DOC-001 was and what replaced it (likely STD-DOC-002). Add note to registry.",
    },
    {
        "id": "CTR-005",
        "type": "Version Inconsistency",
        "sources": "Multiple standards",
        "description": "Multiple standards reference outdated versions of other standards in their cross-reference text: STD-DOC-002 refs STD-DOC-003 v2.1 (current: v2.1.3); STD-DOC-004 refs STD-DOC-002 v2.1 (current: v2.2.0); STD-ENV-002 refs STD-ENV-001 v1.1 (current: v2.0).",
        "impact": "MEDIUM",
        "resolution": "Update all in-body version references to match current versions, or use 'v2.x' style to avoid constant updates.",
    },
    {
        "id": "CTR-006",
        "type": "Severity Self-Contradiction",
        "sources": "STD-ENV-002",
        "description": "STD-ENV-002 declares itself as [C] Critical in header but footer states 'complies with MARKDOWN_STANDARD v2.1 (level [W])'. This implies the document itself is only [W] compliant while claiming [C] governance level.",
        "impact": "LOW",
        "resolution": "Clarify: [C] is governance level, [W] is compliance with markdown formatting. These are different dimensions. Add a note.",
    },
]


# ============================================================
# STYLES
# ============================================================

HEADER_FONT = Font(name="Inter", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Inter", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
INFO_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_body(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border = THIN_BORDER


def severity_fill(level_str):
    if "[C]" in level_str or "Critical" in level_str or "HIGH" in level_str:
        return CRITICAL_FILL
    if "[W]" in level_str or "Warning" in level_str or "MEDIUM" in level_str:
        return WARNING_FILL
    if "[I]" in level_str or "Info" in level_str or "LOW" in level_str:
        return INFO_FILL
    return None


# ============================================================
# SHEET 1: Registry
# ============================================================

def build_registry(ws):
    ws.title = "1-Registry"
    headers = ["STD-ID", "File", "Name", "Version", "Level (self)", "Level (ID system)", "Level (Impl Order)", "Consistency"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for sid, data in STANDARDS.items():
        vals = [
            sid,
            data["file"],
            data["name"],
            data["version"],
            data["level_declared"],
            data["level_in_id_system"],
            data["level_in_impl_order"],
        ]
        # Consistency check
        declared = data["level_declared"].replace(" ", "")
        in_id = data["level_in_id_system"].replace(" ", "")
        in_impl = data["level_in_impl_order"].replace(" ", "")

        if declared == in_id == in_impl:
            consistency = "[OK] Consistent"
        elif in_id == "(not in registry)" or in_impl == "(not in table)":
            consistency = "[FAIL] Missing from registry/table"
        else:
            consistency = "[FAIL] Mismatch"

        vals.append(consistency)

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
        style_body(ws, row, len(headers))

        # Color-code severity columns
        for col_idx in [5, 6, 7]:
            fill = severity_fill(str(ws.cell(row=row, column=col_idx).value))
            if fill:
                ws.cell(row=row, column=col_idx).fill = fill

        # Color consistency
        consist_cell = ws.cell(row=row, column=8)
        if "[OK]" in consistency:
            consist_cell.fill = OK_FILL
        else:
            consist_cell.fill = CRITICAL_FILL

        row += 1

    # Column widths
    widths = [14, 34, 24, 10, 18, 22, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 2: Cross-Reference Matrix
# ============================================================

def build_matrix(ws):
    ws.title = "2-CrossRef Matrix"
    sorted_ids = sorted(STANDARDS.keys())

    # Header row + column
    ws.cell(row=1, column=1, value="Source -> Target")
    for col, sid in enumerate(sorted_ids, 2):
        ws.cell(row=1, column=col, value=sid)
        ws.cell(row=col, column=1, value=sid)
    style_header(ws, 1, len(sorted_ids) + 1)
    for col in range(2, len(sorted_ids) + 2):
        cell = ws.cell(row=col, column=1)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Fill matrix
    for row_idx, src in enumerate(sorted_ids, 2):
        targets = CROSS_REFS.get(src, [])
        for col_idx, tgt in enumerate(sorted_ids, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if tgt in targets:
                cell.value = "X"
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                cell.font = Font(name="Inter", bold=True, color="FFFFFF", size=10)
            elif src == tgt:
                cell.value = "-"
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            else:
                cell.value = ""

    # Check for asymmetric links
    for row_idx, src in enumerate(sorted_ids, 2):
        targets = CROSS_REFS.get(src, [])
        for tgt in targets:
            if tgt in CROSS_REFS and src not in CROSS_REFS[tgt]:
                col_idx = sorted_ids.index(tgt) + 2
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                cell.value = "X!"
                cell.font = Font(name="Inter", bold=True, color="FFFFFF", size=10)

    # Legend
    legend_row = len(sorted_ids) + 4
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Inter", bold=True, size=10)
    ws.cell(row=legend_row + 1, column=1, value="X = bidirectional reference")
    ws.cell(row=legend_row + 1, column=3, value="X! = asymmetric (no back-reference)")
    ws.cell(row=legend_row + 2, column=1, value="Orange cells = missing back-reference (issue)")

    # Widths
    ws.column_dimensions["A"].width = 16
    for col in range(2, len(sorted_ids) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ============================================================
# SHEET 3: Severity Consistency
# ============================================================

def build_severity(ws):
    ws.title = "3-Severity Check"
    headers = ["Issue ID", "Type", "Source", "Section", "Description", "Severity", "Recommendation"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for issue in SEVERITY_ISSUES:
        vals = [issue["id"], issue["type"], issue["source"], issue["section"],
                issue["description"], issue["severity"], issue["recommendation"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
        style_body(ws, row, len(headers))

        sev_cell = ws.cell(row=row, column=6)
        if issue["severity"] == "HIGH":
            sev_cell.fill = CRITICAL_FILL
        elif issue["severity"] == "MEDIUM":
            sev_cell.fill = WARNING_FILL
        else:
            sev_cell.fill = INFO_FILL
        row += 1

    widths = [10, 20, 14, 24, 60, 10, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 4: Broken/Asymmetric References
# ============================================================

def build_refs(ws):
    ws.title = "4-Broken Refs"
    headers = ["Issue ID", "Type", "Source", "Section", "Target", "Detail", "Severity"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for ref in BROKEN_REFS:
        vals = [ref["id"], ref["type"], ref["source"], ref["section"],
                ref.get("target", ""), ref["detail"], ref["severity"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
        style_body(ws, row, len(headers))

        sev_cell = ws.cell(row=row, column=7)
        if ref["severity"] == "HIGH":
            sev_cell.fill = CRITICAL_FILL
        elif ref["severity"] == "MEDIUM":
            sev_cell.fill = WARNING_FILL
        else:
            sev_cell.fill = INFO_FILL
        row += 1

    widths = [10, 24, 14, 20, 24, 60, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 5: Contradictions
# ============================================================

def build_contradictions(ws):
    ws.title = "5-Contradictions"
    headers = ["Issue ID", "Type", "Sources", "Description", "Impact", "Resolution"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for c in CONTRADICTIONS:
        vals = [c["id"], c["type"], c["sources"], c["description"], c["impact"], c["resolution"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
        style_body(ws, row, len(headers))

        impact_cell = ws.cell(row=row, column=5)
        if c["impact"] == "HIGH":
            impact_cell.fill = CRITICAL_FILL
        elif c["impact"] == "MEDIUM":
            impact_cell.fill = WARNING_FILL
        else:
            impact_cell.fill = INFO_FILL
        row += 1

    widths = [10, 24, 28, 60, 10, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 6: Summary
# ============================================================

def build_summary(ws):
    ws.title = "6-Summary"
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="Standards Cross-Test Summary").font = Font(name="Inter", bold=True, size=14, color="1F4E79")

    row = 3
    summary_data = [
        ("Total standards analyzed", "19"),
        ("Total cross-references", str(sum(len(v) for v in CROSS_REFS.values()))),
        ("", ""),
        ("SEVERITY ISSUES", ""),
        ("  High", str(sum(1 for i in SEVERITY_ISSUES if i["severity"] == "HIGH"))),
        ("  Medium", str(sum(1 for i in SEVERITY_ISSUES if i["severity"] == "MEDIUM"))),
        ("  Low", str(sum(1 for i in SEVERITY_ISSUES if i["severity"] == "LOW"))),
        ("", ""),
        ("BROKEN/ASYMMETRIC REFERENCES", ""),
        ("  High", str(sum(1 for r in BROKEN_REFS if r["severity"] == "HIGH"))),
        ("  Medium", str(sum(1 for r in BROKEN_REFS if r["severity"] == "MEDIUM"))),
        ("  Low", str(sum(1 for r in BROKEN_REFS if r["severity"] == "LOW"))),
        ("", ""),
        ("CONTRADICTIONS", ""),
        ("  High impact", str(sum(1 for c in CONTRADICTIONS if c["impact"] == "HIGH"))),
        ("  Medium impact", str(sum(1 for c in CONTRADICTIONS if c["impact"] == "MEDIUM"))),
        ("  Low impact", str(sum(1 for c in CONTRADICTIONS if c["impact"] == "LOW"))),
        ("", ""),
        ("KEY FINDINGS", ""),
        ("1", "STD-DOC-002 [W] vs STD-DOC-003 [C] — same rule, different severity for emoji/unicode"),
        ("2", "STD-A11Y-001 marked DEPRECATED in registry but actively referenced by 3 standards"),
        ("3", "STD-DOC-004 missing from STD-META-001 registry and STD-ARCH-001 table"),
        ("4", "STD-ENV-002 has 3+ missing back-references (most isolated in cross-ref network)"),
        ("5", "Multiple stale version references (v2.1 -> v2.2.0, v1.1 -> v2.0)"),
        ("6", "STD-DOC-001 listed as DEPRECATED in registry but no replacement documented"),
    ]

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = Font(
            name="Inter",
            bold=label.isupper() or label.startswith("KEY") or label in ("1", "2", "3", "4", "5", "6"),
            size=10
        )
        ws.cell(row=row, column=2, value=value).font = Font(name="Inter", size=10)
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 70


# ============================================================
# MAIN
# ============================================================

def main():
    wb = Workbook()

    build_summary(wb.active)
    build_registry(wb.create_sheet())
    build_matrix(wb.create_sheet())
    build_severity(wb.create_sheet())
    build_refs(wb.create_sheet())
    build_contradictions(wb.create_sheet())

    output_path = os.environ.get(
        'REPORT_OUTPUT_PATH',
        os.path.join(REPO_ROOT, 'download', 'standards_cross_test_report.xlsx')
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Report saved to: {output_path}")


if __name__ == "__main__":
    main()
